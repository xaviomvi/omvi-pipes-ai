import asyncio
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Union

from langchain.chat_models.base import BaseChatModel
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
)

from app.models.blocks import (
    Block,
    BlockContainerIndex,
    BlockGroup,
    BlocksContainer,
    BlockType,
    DataFormat,
    GroupType,
    TableMetadata,
)
from app.modules.parsers.excel.prompt_template import (
    prompt,
    row_text_prompt,
    sheet_summary_prompt,
    table_summary_prompt,
)


class ExcelParser:
    def __init__(self, logger) -> None:
        self.logger = logger
        self.workbook = None
        self.file_binary = None

        # Store prompts
        self.sheet_summary_prompt = sheet_summary_prompt
        self.table_summary_prompt = table_summary_prompt
        self.row_text_prompt = row_text_prompt

        # Configure retry parameters
        self.max_retries = 3
        self.min_wait = 1  # seconds
        self.max_wait = 10  # seconds

    async def parse(self, file_binary: bytes, llm: BaseChatModel) -> BlocksContainer:
        """
        Parse Excel file and extract all content including sheets, cells, formulas, etc.

        Returns:
            Dict containing parsed content with structure:
            {
                'sheets': List[Dict],        # List of sheet data
                'metadata': Dict,            # Workbook metadata
                'text_content': str,         # All text content concatenated
                'sheet_names': List[str],    # List of sheet names
                'total_rows': int,           # Total rows across all sheets
                'total_cells': int           # Total cells with content
            }
        """
        try:
            self.file_binary = file_binary
            # Load workbook from binary or file path
            if self.file_binary:
                self.workbook = load_workbook(
                    io.BytesIO(self.file_binary), data_only=True
                )
            else:
                self.workbook = load_workbook(self.file_path, data_only=True)

            return await self.get_blocks_from_workbook(llm)

        except Exception:
            raise
        finally:
            if self.workbook:
                self.workbook.close()

    def _json_default(self, obj) -> str:
        if isinstance(obj, datetime):
            return obj.isoformat()
        return str(obj)

    def _process_sheet(self, sheet) -> Dict[str, List[List[Dict[str, Any]]]]:
        """Process individual sheet and extract cell data"""
        try:
            sheet_data = {"headers": [], "data": []}

            # Extract headers from first row
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
            sheet_data["headers"] = [cell.value for cell in first_row]

            # Start from second row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                row_data = []

                for col_idx, cell in enumerate(row, 1):
                    # Handle merged cells
                    if isinstance(cell, MergedCell):
                        cell_data = {
                            "value": None,  # Merged cells don't contain values
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            # Use utility function instead
                            "column_letter": get_column_letter(col_idx),
                            "coordinate": f"{get_column_letter(col_idx)}{row_idx}",
                            "data_type": "merged",
                            "style": {"font": {}, "fill": {}, "alignment": {}},
                        }
                    else:
                        cell_data = {
                            "value": cell.value,
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            "column_letter": cell.column_letter,
                            "coordinate": cell.coordinate,
                            "data_type": cell.data_type,
                            "style": {
                                "font": {
                                    "bold": cell.font.bold,
                                    "italic": cell.font.italic,
                                    "size": cell.font.size,
                                    "color": (
                                        cell.font.color.rgb if cell.font.color else None
                                    ),
                                },
                                "fill": {
                                    "background_color": (
                                        cell.fill.start_color.rgb
                                        if cell.fill.start_color
                                        else None
                                    )
                                },
                                "alignment": {
                                    "horizontal": cell.alignment.horizontal,
                                    "vertical": cell.alignment.vertical,
                                },
                            },
                        }

                        # Add formula if present
                        if cell.data_type == "f":
                            cell_data["formula"] = cell.value

                    row_data.append(cell_data)

                sheet_data["data"].append(row_data)

            return sheet_data

        except Exception:
            raise

    def find_tables(self, sheet) -> List[Dict[str, Any]]:
        """Find and process all tables in a sheet"""
        try:
            tables = []
            visited_cells = set()  # Track already processed cells

            def get_table(start_row: int, start_col: int) -> Dict[str, Any]:
                """Extract a table starting from (start_row, start_col)."""
                # Find the last column of the table
                max_col = start_col
                for col in range(start_col, sheet.max_column + 1):
                    has_data = False
                    for r in range(start_row, sheet.max_row + 1):
                        cell = sheet.cell(row=r, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_col = col
                            break
                    if not has_data:
                        break

                # Find the last row of the table
                max_row = start_row
                for row in range(start_row, sheet.max_row + 1):
                    has_data = False
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_row = row
                            break
                    if not has_data:
                        break

                # Now process the rectangular table region
                table_data = []
                headers = []

                # Process header row
                header_cells = []
                for col in range(start_col, max_col + 1):
                    cell = sheet.cell(row=start_row, column=col)
                    header_value = self._process_cell(cell, None, start_row, col)
                    header_cells.append(header_value)
                    if cell.value is not None:
                        visited_cells.add((start_row, col))

                # Only consider it a header row if at least one cell has data
                if any(cell["value"] is not None for cell in header_cells):
                    headers = [cell["value"] for cell in header_cells]
                    table_data.append(header_cells)
                else:
                    return {
                        "headers": [],
                        "data": [],
                        "start_row": start_row,
                        "start_col": start_col,
                        "end_row": start_row,
                        "end_col": start_col,
                    }

                # Process data rows within the determined boundaries
                for row in range(start_row + 1, max_row + 1):
                    row_data = []
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        header = (
                            headers[col - start_col]
                            if col - start_col < len(headers)
                            else None
                        )
                        cell_data = self._process_cell(cell, header, row, col)
                        if cell.value is not None:
                            visited_cells.add((row, col))
                        row_data.append(cell_data)
                    table_data.append(row_data)

                return {
                    "headers": headers,
                    "data": table_data[1:] if table_data else [],
                    "start_row": start_row,
                    "start_col": start_col,
                    "end_row": max_row,
                    "end_col": max_col,
                }

            # Find all tables in the sheet
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)

                    # Possible table header detection (assumes headers are text-based)
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and (row, col) not in visited_cells
                    ):
                        table = get_table(row, col)
                        if table["data"]:  # Only add if table has data
                            tables.append(table)

            return tables

        except Exception:
            raise

    def _process_cell(self, cell, header, row, col) -> Dict[str, Any]:
        """Process a single cell and return its data with denormalized merged cell values."""
        try:
            # Check if the cell is a merged cell
            if isinstance(cell, MergedCell):
                # Look for the merged range that contains this cell.
                merged_value = None
                for merged_range in cell.parent.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = cell.parent.cell(
                            row=merged_range.min_row, column=merged_range.min_col
                        )
                        merged_value = top_left_cell.value
                        break

                return {
                    "value": merged_value,  # Use the top-left cell's value
                    "header": header,
                    "row": row,
                    "column": col,
                    "column_letter": get_column_letter(col),
                    "coordinate": f"{get_column_letter(col)}{row}",
                    "data_type": "merged",
                    "style": {"font": {}, "fill": {}, "alignment": {}},
                }

            # If not a merged cell, process normally.
            return {
                "value": cell.value,
                "header": header,
                "row": row,
                "column": col,
                "column_letter": cell.column_letter,
                "coordinate": cell.coordinate,
                "data_type": cell.data_type,
                "style": {
                    "font": {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "color": cell.font.color.rgb if cell.font.color else None,
                    },
                    "fill": {
                        "background_color": (
                            cell.fill.start_color.rgb if cell.fill.start_color else None
                        )
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                    },
                },
            }
        except Exception:
            raise

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=1, max=10),
        before_sleep=lambda retry_state: retry_state.args[0].logger.warning(
            f"Retrying LLM call after error. Attempt {retry_state.attempt_number}"
        ),
    )
    async def _call_llm(self, messages) -> Union[str, dict, list]:
        """Wrapper for LLM calls with retry logic"""
        return await self.llm.ainvoke(messages)

    async def get_tables_in_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get all tables in a specific sheet"""
        try:
            if not self.workbook:
                self.parse()

            if sheet_name not in self.workbook.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return []

            sheet = self.workbook[sheet_name]
            tables = self.find_tables(sheet)

            # Prepare context for LLM with all tables
            tables_context = []
            for idx, table in enumerate(tables, 1):
                table_data = [[cell["value"] for cell in row] for row in table["data"][:10]]
                tables_context.append(f"Table {idx}:\n{table_data}")

            # Process each table with LLM
            processed_tables = []
            for idx, table in enumerate(tables, 1):
                table_data = [[cell["value"] for cell in row] for row in table["data"][:10]]

                # Use prompt from prompt_template.py
                formatted_prompt = prompt.format(
                    table_data=table_data,
                    tables_context=tables_context,
                    start_row=table["start_row"],
                    start_col=table["start_col"],
                    end_row=table["end_row"],
                    end_col=table["end_col"],
                    num_columns=len(table["data"][0]) if table["data"] else 0,
                )

                # Get LLM response with retry
                messages = [
                    {
                        "role": "system",
                        "content": "You are a data analysis expert. Respond with only the list of headers.",
                    },
                    {"role": "user", "content": formatted_prompt},
                ]
                response = await self._call_llm(messages)
                if '</think>' in response.content:
                    response.content = response.content.split('</think>')[-1]

                try:
                    # Parse LLM response to get headers
                    new_headers = [
                        h.strip() for h in response.content.strip().split(",")
                    ]

                    # Ensure we have the right number of headers
                    if len(new_headers) != len(table["data"][0]):
                        new_headers = table["headers"]

                    # Reconstruct table with new headers
                    new_table = {
                        "headers": new_headers,
                        "data": table["data"],
                        "start_row": table["start_row"],
                        "start_col": table["start_col"],
                        "end_row": table["end_row"],
                        "end_col": table["end_col"],
                    }

                    # Update cell header references in the data
                    for row in new_table["data"]:
                        for i, cell in enumerate(row):
                            cell["header"] = (
                                new_headers[i] if i < len(new_headers) else None
                            )

                    processed_tables.append(new_table)

                except Exception:
                    # Fall back to original table
                    processed_tables.append(table)

            return processed_tables

        except Exception:
            raise

    async def get_table_summary(self, table: Dict[str, Any]) -> str:
        """Get a natural language summary of a specific table"""
        try:
            # Prepare sample data
            sample_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in table["data"][:3]  # Use first 3 rows as sample
            ]

            # Get summary from LLM with retry
            messages = self.table_summary_prompt.format_messages(
                headers=table["headers"], sample_data=json.dumps(sample_data, indent=2)
            )
            response = await self._call_llm(messages)
            if '</think>' in response.content:
                response.content = response.content.split('</think>')[-1]
            return response.content

        except Exception:
            raise

    async def get_rows_text(
        self, rows: List[List[Dict[str, Any]]], table_summary: str
    ) -> List[str]:
        """Convert multiple rows into natural language text using context from summaries in a single prompt"""
        try:
            # Prepare rows data
            rows_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in rows
            ]

            # Get natural language text from LLM with retry
            messages = self.row_text_prompt.format_messages(
                table_summary=table_summary, rows_data=json.dumps(rows_data, indent=2)
            )

            response = await self._call_llm(messages)
            if '</think>' in response.content:
                response.content = response.content.split('</think>')[-1]
            # Try to extract JSON array from response
            try:
                # First try direct JSON parsing
                return json.loads(response.content)
            except json.JSONDecodeError:
                # If that fails, try to find and parse a JSON array in the response
                content = response.content
                # Look for array between [ and ]
                start = content.find("[")
                end = content.rfind("]")
                if start != -1 and end != -1:
                    try:
                        return json.loads(content[start : end + 1])
                    except json.JSONDecodeError:
                        # If still can't parse, return response as single-item array
                        return [content]
                else:
                    # If no array found, return response as single-item array
                    return [content]

        except Exception:
            raise

    async def process_sheet_with_summaries(
        self, llm, sheet_name: str
    ) -> Dict[str, Any]:
        """Process a sheet and generate all summaries and row texts"""
        self.llm = llm

        if sheet_name not in self.workbook.sheetnames:
            self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return None

        # Get tables in the sheet
        tables = await self.get_tables_in_sheet(sheet_name)

        # Process each table
        processed_tables = []
        for table in tables:
            # Get table summary
            table_summary = await self.get_table_summary(table)

            # Process rows in batches of 50 in parallel
            processed_rows = []
            batch_size = 50

            # Create batches
            batches = []
            for i in range(0, len(table["data"]), batch_size):
                batch = table["data"][i : i + batch_size]
                batches.append((i, batch))  # Store start index and batch data

            # Limit parallel processing to at most 10 concurrent batches
            semaphore = asyncio.Semaphore(10)

            async def limited_get_rows_text(batch) -> List[str]:
                async with semaphore:
                    return await self.get_rows_text(batch, table_summary)

            # Create throttled tasks for all batches
            batch_tasks = []
            for start_idx, batch in batches:
                task = limited_get_rows_text(batch)
                batch_tasks.append((start_idx, batch, task))

            # Wait for all batches to complete (max 10 running concurrently)
            task_results = await asyncio.gather(*[task for _, _, task in batch_tasks])

            # Combine results with their metadata and process
            for i, (start_idx, batch, _) in enumerate(batch_tasks):
                row_texts = task_results[i]

                # Add processed rows to results
                for row, row_text in zip(batch, row_texts):
                    processed_rows.append(
                        {
                            "raw_data": {cell["header"]: cell["value"] for cell in row},
                            "natural_language_text": row_text,
                            "row_num": row[0]["row"],  # Include row number
                        }
                    )

            processed_tables.append(
                {
                    "headers": table["headers"],
                    "summary": table_summary,
                    "rows": processed_rows,
                    "location": {
                        "start_row": table["start_row"],
                        "start_col": table["start_col"],
                        "end_row": table["end_row"],
                        "end_col": table["end_col"],
                    },
                }
            )

        return {"sheet_name": sheet_name, "tables": processed_tables}

    async def get_blocks_from_workbook(self, llm) -> BlocksContainer:
        """Build a BlocksContainer with SHEET and TABLE groups and TABLE_ROW blocks.

        Mirrors the CSV blocks structure, but nests tables under sheet groups.
        """
        blocks: List[Block] = []
        block_groups: List[BlockGroup] = []

        # Iterate sheets and build hierarchy
        for sheet_idx, sheet_name in enumerate(self.workbook.sheetnames, 1):
            sheet_result = await self.process_sheet_with_summaries(llm, sheet_name)
            if sheet_result is None:
                continue

            # Create SHEET group
            sheet_group_index = len(block_groups)
            sheet_group_children: List[BlockContainerIndex] = []
            sheet_group = BlockGroup(
                index=sheet_group_index,
                name=sheet_result["sheet_name"],
                type=GroupType.SHEET,
                parent_index=None,
                description=None,
                table_metadata=None,
                data={
                    "sheet_name": sheet_result["sheet_name"],
                    "table_count": len(sheet_result["tables"]),
                },
                format=DataFormat.JSON,
            )
            block_groups.append(sheet_group)

            # Add TABLE groups under this sheet
            for table in sheet_result["tables"]:
                table_group_index = len(block_groups)

                headers = table.get("headers", [])
                rows = table.get("rows", [])

                table_group_children: List[BlockContainerIndex] = []
                table_markdown = self.to_markdown(headers, rows)
                table_group = BlockGroup(
                    index=table_group_index,
                    name=None,
                    type=GroupType.TABLE,
                    parent_index=sheet_group_index,
                    description=None,
                    source_group_id=None,
                    table_metadata=TableMetadata(
                        num_of_rows=len(rows),
                        num_of_cols=len(headers) if headers else (len(rows[0]["raw_data"]) if rows else 0),
                    ),
                    data={
                        "table_summary": table.get("summary", ""),
                        "column_headers": headers,
                        "table_markdown": table_markdown,
                        "sheet_number": sheet_idx,
                        "sheet_name": sheet_name,
                    },
                    format=DataFormat.JSON,
                )
                block_groups.append(table_group)
                sheet_group_children.append(BlockContainerIndex(block_group_index=table_group_index))

                # Create TABLE_ROW blocks under this table
                for i, row in enumerate(rows):
                    block_index = len(blocks)
                    row_data = row.get("raw_data", {})
                    blocks.append(
                        Block(
                            index=block_index,
                            type=BlockType.TABLE_ROW,
                            format=DataFormat.JSON,
                            data={
                                "row_natural_language_text": row.get("natural_language_text", ""),
                                "row_number": int(row.get("row_num") or (i + 1)),
                                "row": json.dumps(row_data, default=self._json_default),
                                "sheet_number": sheet_idx,
                                "sheet_name": sheet_name,
                            },
                            parent_index=table_group_index,
                        )
                    )
                    table_group_children.append(BlockContainerIndex(block_index=block_index))

                # attach table children
                block_groups[table_group_index].children = table_group_children

            # attach sheet children (its tables)
            block_groups[sheet_group_index].children = sheet_group_children

        return BlocksContainer(blocks=blocks, block_groups=block_groups)

    def to_markdown(self, headers: List[str], rows: List[Dict[str, Any]]) -> str:
        """
        Convert CSV data to markdown table format.
        Args:
            data: List of dictionaries from read_stream() method
        Returns:
            String containing markdown formatted table
        """
        if not headers and not rows:
            return ""

        # Get headers from the first row
        headers = list(headers)

        # Start building the markdown table
        markdown_lines = []

        # Add header row
        header_row = "| " + " | ".join(str(header) for header in headers) + " |"
        markdown_lines.append(header_row)

        # Add separator row
        separator_row = "|" + "|".join(" --- " for _ in headers) + "|"
        markdown_lines.append(separator_row)
        data = []
        for row in rows:
            data.append(row.get("raw_data", {}))
        # Add data rows
        for row in data:
            # Handle None values and convert to string, escape pipe characters
            formatted_values = []
            for header in headers:
                value = row.get(header, "")
                if value is None:
                    value = ""
                # Escape pipe characters and convert to string
                value_str = str(value).replace("|", "\\|")
                formatted_values.append(value_str)

            data_row = "| " + " | ".join(formatted_values) + " |"
            markdown_lines.append(data_row)

        return "\n".join(markdown_lines)

