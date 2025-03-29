from openpyxl import load_workbook
from typing import Dict, List, Any
import io
import os
from app.utils.logger import logger
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from app.core.llm_service import LLMFactory
from app.modules.parsers.excel.prompt_template import prompt, sheet_summary_prompt, table_summary_prompt, row_text_prompt
import json
from datetime import datetime


class ExcelParser:
    def __init__(self, llm_config):
        self.workbook = None
        self.file_binary = None

        self.llm = LLMFactory.create_llm(llm_config)

        # Store prompts
        self.sheet_summary_prompt = sheet_summary_prompt
        self.table_summary_prompt = table_summary_prompt
        self.row_text_prompt = row_text_prompt

    def parse(self, file_binary: bytes) -> Dict[str, Any]:
        """
        Parse Excel file and extract all content including sheets, cells, formulas, etc.

        Returns:
            Dict containing parsed content with structure:
            {
                'sheets': List[Dict],        # List of sheet data
                'metadata': Dict,            # Workbook metadata
                'text_content': str,         # All text content concatenated
                'sheet_names': List[str],    # List of sheet names
                'total_rows': int,           # Total rows across all sheets
                'total_cells': int           # Total cells with content
            }
        """
        try:
            self.file_binary = file_binary
            # Load workbook from binary or file path
            if self.file_binary:
                self.workbook = load_workbook(
                    io.BytesIO(self.file_binary), data_only=True)
            else:
                self.workbook = load_workbook(self.file_path, data_only=True)

            sheets_data = []
            total_rows = 0
            total_cells = 0
            all_text = []

            # Process each sheet
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet_data = self._process_sheet(sheet)

                sheets_data.append({
                    'name': sheet_name,
                    'data': sheet_data['data'],
                    'headers': sheet_data['headers'],
                    'row_count': sheet.max_row,
                    'column_count': sheet.max_column,
                    'merged_cells': [str(merged_range) for merged_range in sheet.merged_cells.ranges]
                })

                total_rows += sheet.max_row
                total_cells += sum(1 for row in sheet_data['data']
                                   for cell in row if cell['value'])

                all_text.extend([
                    str(cell['value']) for row in sheet_data['data']
                    for cell in row if cell['value'] is not None
                ])

            # Prepare metadata
            metadata = {
                'creator': self.workbook.properties.creator,
                'created': self.workbook.properties.created.isoformat() if self.workbook.properties.created else None,
                'modified': self.workbook.properties.modified.isoformat() if self.workbook.properties.modified else None,
                'last_modified_by': self.workbook.properties.lastModifiedBy,
                'sheet_count': len(self.workbook.sheetnames)
            }

            return {
                'sheets': sheets_data,
                'metadata': metadata,
                'text_content': '\n'.join(all_text),
                'sheet_names': self.workbook.sheetnames,
                'total_rows': total_rows,
                'total_cells': total_cells
            }

        except Exception as e:
            logger.error(f"❌ Error parsing Excel file: {str(e)}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()

    def _process_sheet(self, sheet) -> Dict[str, List[List[Dict[str, Any]]]]:
        """Process individual sheet and extract cell data"""
        try:
            logger.debug(f"Processing sheet: {sheet.title}")
            sheet_data = {'headers': [], 'data': []}

            # Extract headers from first row
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
            sheet_data['headers'] = [cell.value for cell in first_row]

            # Start from second row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                row_data = []

                for col_idx, cell in enumerate(row, 1):
                    # Handle merged cells
                    if isinstance(cell, MergedCell):
                        cell_data = {
                            'value': None,  # Merged cells don't contain values
                            'header': sheet_data['headers'][col_idx-1] if col_idx-1 < len(sheet_data['headers']) else None,
                            'row': row_idx,
                            'column': col_idx,
                            # Use utility function instead
                            'column_letter': get_column_letter(col_idx),
                            'coordinate': f"{get_column_letter(col_idx)}{row_idx}",
                            'data_type': 'merged',
                            'style': {
                                'font': {},
                                'fill': {},
                                'alignment': {}
                            }
                        }
                    else:
                        cell_data = {
                            'value': cell.value,
                            'header': sheet_data['headers'][col_idx-1] if col_idx-1 < len(sheet_data['headers']) else None,
                            'row': row_idx,
                            'column': col_idx,
                            'column_letter': cell.column_letter,
                            'coordinate': cell.coordinate,
                            'data_type': cell.data_type,
                            'style': {
                                'font': {
                                    'bold': cell.font.bold,
                                    'italic': cell.font.italic,
                                    'size': cell.font.size,
                                    'color': cell.font.color.rgb if cell.font.color else None
                                },
                                'fill': {
                                    'background_color': cell.fill.start_color.rgb if cell.fill.start_color else None
                                },
                                'alignment': {
                                    'horizontal': cell.alignment.horizontal,
                                    'vertical': cell.alignment.vertical
                                }
                            }
                        }

                        # Add formula if present
                        if cell.data_type == 'f':
                            cell_data['formula'] = cell.value

                    row_data.append(cell_data)

                sheet_data['data'].append(row_data)

            logger.debug(f"Processed sheet: {sheet.title}")
            return sheet_data

        except Exception as e:
            logger.error(f"❌ Error processing sheet: {str(e)}")
            raise

    def get_sheet_data(self, sheet_name: str = None) -> Dict[str, Any]:
        """
        Get data for a specific sheet or all sheets

        Args:
            sheet_name: Name of the sheet to get data from. If None, returns all sheets.

        Returns:
            Dictionary containing sheet data with headers and rows
        """
        try:
            logger.debug(f"Getting sheet data for: {sheet_name}")
            if not self.workbook:
                self.parse()

            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found in workbook")
                sheet = self.workbook[sheet_name]
                return self._process_sheet(sheet)

            return {name: self._process_sheet(self.workbook[name]) for name in self.workbook.sheetnames}

        except Exception as e:
            logger.error(f"❌ Error getting sheet data: {str(e)}")
            raise

    def find_tables(self, sheet) -> List[Dict[str, Any]]:
        """Find and process all tables in a sheet"""
        try:
            logger.debug(f"Finding tables in sheet: {sheet.title}")
            tables = []
            visited_cells = set()  # Track already processed cells

            def get_table(start_row, start_col):
                """Extract a table starting from (start_row, start_col)."""
                table_data = []
                headers = []
                row = start_row

                # Get the first row as headers
                first_row = []
                col = start_col
                while col <= sheet.max_column:
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        visited_cells.add((row, col))
                        first_row.append(
                            self._process_cell(cell, None, row, col))
                    else:
                        break
                    col += 1

                if first_row:
                    headers = [cell['value'] for cell in first_row]
                    table_data.append(first_row)

                # Process remaining rows
                row += 1
                while row <= sheet.max_row:
                    row_data = []
                    empty_row = True
                    col = start_col

                    while col <= sheet.max_column:
                        cell = sheet.cell(row=row, column=col)
                        header = headers[col - start_col] if col - \
                            start_col < len(headers) else None

                        if cell.value is not None:
                            empty_row = False
                            visited_cells.add((row, col))
                            row_data.append(self._process_cell(
                                cell, header, row, col))
                        else:
                            break
                        col += 1

                    if empty_row:
                        break
                    if row_data:
                        table_data.append(row_data)
                    row += 1

                return {
                    'headers': headers,
                    # Exclude headers from data
                    'data': table_data[1:] if table_data else [],
                    'start_row': start_row,
                    'start_col': start_col,
                    'end_row': row - 1,
                    'end_col': col - 1
                }

            # Find all tables in the sheet
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)

                    # Possible table header detection (assumes headers are text-based)
                    if cell.value and isinstance(cell.value, str) and (row, col) not in visited_cells:
                        table = get_table(row, col)
                        if table['data']:  # Only add if table has data
                            tables.append(table)

            logger.debug(f"Found {len(tables)} tables in sheet: {sheet.title}")
            return tables

        except Exception as e:
            logger.error(f"❌ Error finding tables: {str(e)}")
            raise

    def _process_cell(self, cell, header, row, col):
        """Process a single cell and return its data"""

        try:
            if isinstance(cell, MergedCell):
                return {
                    'value': None,
                    'header': header,
                    'row': row,
                    'column': col,
                    'column_letter': get_column_letter(col),
                    'coordinate': f"{get_column_letter(col)}{row}",
                    'data_type': 'merged',
                    'style': {
                        'font': {},
                        'fill': {},
                        'alignment': {}
                    }
                }

            return {
                'value': cell.value,
                'header': header,
                'row': row,
                'column': col,
                'column_letter': cell.column_letter,
                'coordinate': cell.coordinate,
                'data_type': cell.data_type,
                'style': {
                    'font': {
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'size': cell.font.size,
                        'color': cell.font.color.rgb if cell.font.color else None
                    },
                    'fill': {
                        'background_color': cell.fill.start_color.rgb if cell.fill.start_color else None
                    },
                    'alignment': {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical
                    }
                }
            }
        except Exception as e:
            logger.error(f"❌ Error processing cell: {str(e)}")
            raise

    async def get_tables_in_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get all tables in a specific sheet"""
        try:
            logger.debug(f"Getting tables in sheet: {sheet_name}")
            if not self.workbook:
                self.parse()

            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            sheet = self.workbook[sheet_name]
            tables = self.find_tables(sheet)

            # Prepare context for LLM with all tables
            tables_context = []
            for idx, table in enumerate(tables, 1):
                table_data = [
                    [cell['value'] for cell in row]
                    for row in table['data']
                ]
                tables_context.append(f"Table {idx}:\n{table_data}")

            # Process each table with LLM
            processed_tables = []
            for idx, table in enumerate(tables, 1):
                table_data = [
                    [cell['value'] for cell in row]
                    for row in table['data']
                ]

                # Use prompt from prompt_template.py
                formatted_prompt = prompt.format(
                    table_data=table_data,
                    tables_context=tables_context,
                    start_row=table['start_row'],
                    start_col=table['start_col'],
                    end_row=table['end_row'],
                    end_col=table['end_col'],
                    num_columns=len(table['data'][0]) if table['data'] else 0
                )

                # Get LLM response
                messages = [
                    {"role": "system", "content": "You are a data analysis expert. Respond with only the list of headers."},
                    {"role": "user", "content": formatted_prompt}
                ]
                response = await self.llm.ainvoke(messages)

                try:
                    # Parse LLM response to get headers
                    new_headers = [h.strip()
                                   for h in response.content.strip().split(',')]

                    # Ensure we have the right number of headers
                    if len(new_headers) != len(table['data'][0]):
                        logger.warning(f"""LLM generated incorrect number of headers for table {
                                       idx}. Falling back to original headers.""")
                        new_headers = table['headers']

                    # Reconstruct table with new headers
                    new_table = {
                        'headers': new_headers,
                        'data': table['data'],
                        'start_row': table['start_row'],
                        'start_col': table['start_col'],
                        'end_row': table['end_row'],
                        'end_col': table['end_col']
                    }

                    # Update cell header references in the data
                    for row in new_table['data']:
                        for i, cell in enumerate(row):
                            cell['header'] = new_headers[i] if i < len(
                                new_headers) else None

                    processed_tables.append(new_table)

                except Exception as e:
                    logger.error(f"""Error processing headers for table {
                                 idx}: {str(e)}""")
                    # Fall back to original table
                    processed_tables.append(table)

            return processed_tables

        except Exception as e:
            logger.error(f"❌ Error getting tables in sheet: {str(e)}")
            raise

    async def get_sheet_summary(self, sheet_name: str, tables: List[Dict[str, Any]]) -> str:
        """Get a natural language summary of all tables in a sheet"""
        try:
            logger.debug(f"Getting sheet summary for: {sheet_name}")
            # Prepare tables data for the prompt
            tables_data = []
            for idx, table in enumerate(tables, 1):
                sample_data = [
                    {cell['header']: (cell['value'].isoformat() if isinstance(cell['value'], datetime) else cell['value'])
                     for cell in row}
                    for row in table['data'][:3]  # Use first 3 rows as sample
                ]
                tables_data.append(f"Table {idx}:\nHeaders: {table['headers']}\n"
                                   f"Sample data:\n{json.dumps(sample_data, indent=2)}")

            # Get summary from LLM
            messages = self.sheet_summary_prompt.format_messages(
                sheet_name=sheet_name,
                tables_data="\n\n".join(tables_data)
            )
            response = await self.llm.ainvoke(messages)
            return response.content

        except Exception as e:
            logger.error(f"❌ Error getting sheet summary: {str(e)}")
            raise

    async def get_table_summary(self, table: Dict[str, Any]) -> str:
        """Get a natural language summary of a specific table"""
        try:
            logger.debug(f"Getting table summary for: {table['headers']}")
            # Prepare sample data
            sample_data = [
                {cell['header']: (cell['value'].isoformat() if isinstance(cell['value'], datetime) else cell['value'])
                 for cell in row}
                for row in table['data'][:3]  # Use first 3 rows as sample
            ]

            # Get summary from LLM
            messages = self.table_summary_prompt.format_messages(
                headers=table['headers'],
                sample_data=json.dumps(sample_data, indent=2)
            )
            response = await self.llm.ainvoke(messages)
            return response.content

        except Exception as e:
            logger.error(f"❌ Error getting table summary: {str(e)}")
            raise

    async def get_rows_text(self, rows: List[List[Dict[str, Any]]], sheet_summary: str, table_summary: str) -> List[str]:
        """Convert multiple rows into natural language text using context from summaries in a single prompt"""
        try:
            logger.debug(f"Getting rows text for: {rows}")
            # Prepare rows data
            rows_data = [
                {cell['header']: (cell['value'].isoformat() if isinstance(cell['value'], datetime) else cell['value'])
                 for cell in row}
                for row in rows
            ]

            # Get natural language text from LLM for all rows
            messages = self.row_text_prompt.format_messages(
                sheet_summary=sheet_summary,
                table_summary=table_summary,
                rows_data=json.dumps(rows_data, indent=2)
            )

            response = await self.llm.ainvoke(messages)

            # Try to extract JSON array from response
            try:
                # First try direct JSON parsing
                return json.loads(response.content)
            except json.JSONDecodeError:
                # If that fails, try to find and parse a JSON array in the response
                content = response.content
                # Look for array between [ and ]
                start = content.find('[')
                end = content.rfind(']')
                if start != -1 and end != -1:
                    try:
                        return json.loads(content[start:end+1])
                    except json.JSONDecodeError:
                        # If still can't parse, return response as single-item array
                        return [content]
                else:
                    # If no array found, return response as single-item array
                    return [content]

        except Exception as e:
            logger.error(f"❌ Error getting rows text: {str(e)}")
            raise

    async def process_sheet_with_summaries(self, sheet_name: str) -> Dict[str, Any]:
        """Process a sheet and generate all summaries and row texts"""
        if not self.workbook:
            self.parse()

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        # Get tables in the sheet
        tables = await self.get_tables_in_sheet(sheet_name)

        # Get sheet-level summary
        sheet_summary = await self.get_sheet_summary(sheet_name, tables)

        # Process each table
        processed_tables = []
        for table in tables:
            # Get table summary
            table_summary = await self.get_table_summary(table)

            # Process rows in batches of 20
            processed_rows = []
            batch_size = 20

            for i in range(0, len(table['data']), batch_size):
                batch = table['data'][i:i + batch_size]
                row_texts = await self.get_rows_text(batch, sheet_summary, table_summary)
                

                # Add processed rows to results
                for row, row_text in zip(batch, row_texts):
                    processed_rows.append({
                        'raw_data': {cell['header']: cell['value'] for cell in row},
                        'natural_language_text': row_text,
                        'row_num': row[0]['row']  # Include row number
                    })

            processed_tables.append({
                'headers': table['headers'],
                'summary': table_summary,
                'rows': processed_rows,
                'location': {
                    'start_row': table['start_row'],
                    'start_col': table['start_col'],
                    'end_row': table['end_row'],
                    'end_col': table['end_col']
                }
            })

        return {
            'sheet_name': sheet_name,
            'sheet_summary': sheet_summary,
            'tables': processed_tables
        }


async def main():
    """Test function to demonstrate Excel parsing with summaries"""
    test_file = "modules/parsers/excel/test4.xlsx"

    try:
        parser = ExcelParser(file_path=test_file)
        parsed_data = parser.parse()
        print("\nAvailable sheets:", parsed_data['sheet_names'])

        # Process each sheet
        for sheet_name in parsed_data['sheet_names']:
            print(f"\n{'='*50}")
            print(f"Processing sheet: {sheet_name}")
            print(f"{'='*50}")

            sheet_data = await parser.process_sheet_with_summaries(sheet_name)

            print("\nSheet Summary:")
            print(sheet_data['sheet_summary'])

            for idx, table in enumerate(sheet_data['tables'], 1):
                print(f"\nTable {idx} Summary:")
                print(table['summary'])

                print("\nSample row texts:")
                for row in table['rows'][:2]:  # Show first 2 rows
                    print(f"\nRaw data: {row['raw_data']}")
                    print(f"Natural text: {row['natural_language_text']}")

                if len(table['rows']) > 2:
                    print("...")

    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
