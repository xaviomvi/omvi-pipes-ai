from openpyxl import load_workbook


def find_tables(sheet):
    tables = []
    visited_cells = set()  # Track already processed cells

    def get_table(start_row, start_col):
        """Extract a table starting from (start_row, start_col)."""
        table_data = []
        row = start_row

        # Expand downwards
        while row <= sheet.max_row:
            row_values = []
            empty_row = True
            col = start_col

            # Expand rightwards
            while col <= sheet.max_column:
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    empty_row = False
                    visited_cells.add((row, col))
                    row_values.append(cell.value)
                else:
                    break  # Stop at the first empty cell in a row
                col += 1

            if empty_row:
                break  # Stop if we reach an empty row
            table_data.append(row_values)
            row += 1

        return table_data

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)

            # Possible table header detection (assumes headers are text-based)
            if cell.value and isinstance(cell.value, str) and (row, col) not in visited_cells:
                table = get_table(row, col)
                tables.append(table)

    return tables


# Load the workbook and select the specified sheet
file_path = "modules/parsers/excel/test4.xlsx"
wb = load_workbook(file_path)
sheet_name = "Level 2- Medium"

if sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Extract tables
    tables = find_tables(sheet)

    # Print extracted tables
    for idx, table in enumerate(tables):
        print(f"Table {idx+1}:")
        for row in table:
            print(row)
        print("-" * 40)
else:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
